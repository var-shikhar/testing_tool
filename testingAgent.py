"""

V1.0 
Base version of test agent that works

22 Aug 25
1. Update the system prompt to ensure that there is no repeated call to browser_snapshot since every toolcall returns latest browser snapshot
    - This did not work so still using old System message. So it calls snapshot before every call 
2. Handeled reconnecting to MCP servers by calling initialize_mcp_manager(). First time when called it cleans up old instance of the same MCP server and next time it connects

12 Sept 24
1. Handling screenshot was broken. Now handled it again. I think the format in which playwrite was sending image data got changed.

2. Need to improve tracablity of calls on this file. So we need to look at the flow from run_test_suite_from_web function.
So basically I need
a. What was the user input
b. What input was sent to the LLM
c. What was the output from the LLM
d. What tool call was made with what parameters
e. What was the output from the tool
f. What was the final result

All of this needs to be written clearly well formatted in a file starting from initial user request. 

The file needs to be written in a folder name debuglogs and each file needs to have name starting with the test case and datetime appended to it

"""
import os
import json
import asyncio
import openai
import google.generativeai as genai
from mcp.client.sse import sse_client
from mcp import ClientSession
import logging
from contextlib import AsyncExitStack
import copy
from typing import Any, Dict, List, Optional
from collections import namedtuple
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
import base64
import datetime
import re
from dotenv import load_dotenv
from chat_processor import initialize_mcp_manager

# --- New imports for web integration ---
from fastapi import WebSocket
import traceback

# Configure logging (remains useful for server-side debugging)
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(name)s - %(message)s')
logger = logging.getLogger(__name__)

# Special logger for capturing the run dialogue to a file
run_logger = logging.getLogger('RunLogger')
run_logger.setLevel(logging.INFO)
# This logger will be configured with a FileHandler dynamically during web-based test runs.

# New logger for detailed debug tracing of each test case
debug_logger = logging.getLogger('DebugLogger')
debug_logger.setLevel(logging.INFO)
# This logger's handler is managed by the run_test_suite_from_web function.

# Suppress noisy logs from libraries
logging.getLogger("openai").setLevel(logging.WARNING)
logging.getLogger("httpcore").setLevel(logging.WARNING)
logging.getLogger("httpx").setLevel(logging.WARNING)
logging.getLogger("mcp.client.session").setLevel(logging.INFO)
logging.getLogger('google.generativeai').setLevel(logging.WARNING)

# These will be set dynamically in the web-based runner
SCREENSHOT_DIR = './' 

# MCP_ENDPOINTS = {
#     "memory": 'http://localhost:8126/sse',
#     "browser": "http://localhost:8127/sse",
# }

# Load environment variables once at the start
load_dotenv()
MCP_ENDPOINTS = json.loads(os.getenv("MCP_ENDPOINTS_JSON"))
LLM_PROVIDER = os.getenv("LLM_PROVIDER", "openai").lower() # openai or google
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY") # Already used by OpenAI client
OPENAI_MODEL = os.getenv("OPENAI_MODEL", "gpt-5-mini")
BASE_URL = os.getenv("BASE_URL")


# GOOGLE_MODEL = os.getenv("GOOGLE_MODEL", "gemini-2.5-flash-preview-05-20")

MAX_TURNS = int(os.getenv("MAX_TURNS", 25))
SNAPSHOT_LOG_MAX_LEN = 2000
# New constant for truncating old browser tool call results in the history
OLD_BROWSER_SNAPSHOT_TRUNCATION_LENGTH = 200

SYSTEM_PROMPT = """
You are an advanced agent capable of controlling various aspects of a system,
including a web browser, filesystem, and potentially other tools like memory analysis or chart generation.
Your goal is to fulfill the user's request by interacting with the appropriate tools.

Tools are namespaced by their category. For example:
- Browser tools start with `browser_` (e.g., `browser_browser_snapshot`, `browser_browser_click`).
- Memory tools (if available) might start with `memory_` (e.g., `memory_memory_get_usage`).

Tools usage Info:
Memory Tool: Memory works based on Key value, You store values based on keys and can retrive values based on keys. If you pass Key as blank it will return all values. While describing keys users may use many synonyms words like tags, keyword etc
Browser Tool: It has feature to take screeshot as well. When successful, it saves the screenshot in a folder and returns the path of image. It needs to be called only with one argument "filename": "screenshot.png".  In case it gives "Screenshot executed, but no image data found." we need to try again to take screenshot. 

Follow these steps:
1.  **Analyze the Request:** Understand what the user wants to achieve and which category of tools is most relevant.
2.  **Assess the Current State (Browser):** If interacting with the browser and you don't know the current page structure or need to find specific elements, use the `browser_snapshot` tool. This tool returns the accessibility tree and interactable elements with their `mcp_id`s.
3.  **Identify Target Elements (Browser):** Carefully examine the `browser_browser_snapshot` output to find the `mcp_id` of the element(s) you need to interact with.
4.  **Execute Action:** Use the appropriate tool (e.g., `browser_browser_click`, `browser_browser_type`) with the specific arguments required.
    *   For browser actions requiring an `mcp_id`, ensure you have it from a recent `browser_browser_snapshot`.
    *   The `browser_browser_type` tool requires arguments like `{"mcp_id": "element-id-from-snapshot", "text": "text to type"}` and every call to browser tool needs to be followed by updated snapshot request. If we need to type into multiple fields we cannot type multiple based on same snapshot as snapshot udpates after each call. So we need to execute one action and then get snapshot and then execute next action.
5.  **Confirm or Continue:** After an action, assess if the goal is complete. If not, take another snapshot (if needed for browser tasks), or perform the next action. Provide a final confirmation to the user.
6.  **Error Handling:** If a tool call returns an error, analyze the error message, potentially take another snapshot (for browser), try a different approach, or inform the user.
7. This is running in autorun mode, so use your best judgement to proceed to next step, do not ask user for inputs.
8. When in doubt about current state and what elements are available on browser screen take a fresh browser_browser_snapshot.
9. **NEVER** use browser navigate unless explicitly asked by user. 
10. If a field is diabled, then we do not need to try to fill it. 
11. **DO NOT** perform any action that has not been explicitly asked. For example do not click on buttons unless user instructions clearly asks for it. 
"""

MAX_TOOL_RESULT_LEN_IN_HISTORY = 500

ACTIONS_REQUIRING_ELEMENT_INTERACTION = {
    "browser_click",
    "browser_hover",
    "browser_type",
    "browser_select_option",
}

openai_client_instance: Optional[openai.AsyncOpenAI] = None

def get_openai_client() -> openai.AsyncOpenAI:
    global openai_client_instance
    if openai_client_instance is None:
        if not OPENAI_API_KEY:
            logger.error("OPENAI_API_KEY not found. OpenAI client cannot be initialized.")
            raise ValueError("OPENAI_API_KEY is not set.")
        openai_client_instance = openai.AsyncOpenAI(api_key=OPENAI_API_KEY, base_url=BASE_URL)
    return openai_client_instance

async def _get_llm_summary_of_turn(context_summary: str, tool_name: str, tool_result: str) -> str:
    if "[TOOL ERROR]" in tool_result:
        return f"The agent tried to use tool '{tool_name}' but it failed with an error."

    if tool_name in ["browser_navigate", "browser_click", "browser_type"]:
        return f"The agent successfully used the '{tool_name}' tool as part of its plan."

    if len(tool_result) > 1000:
        tool_result = tool_result[:1000] + "... (truncated)"
    
    summary_prompt = f"""
Your job is to analyse the Overall goal, tool call and tool call result related data provided below. 
Extract the data from tool call result  Result that would be important for acheiving overall goal.
Prepare a summary that needs to contain tool name and the important data from result that would be needed to accomplish overall goal
Overall Goal: "{context_summary}"
Action Taken: The agent used the tool '{tool_name}'.
Result: {tool_result}
"""
    try:
        client = get_openai_client()
        response = client.chat.completions.create(
            model='gpt-oss:latest',
            messages=[{"role": "user", "content": summary_prompt}],
        )
        return response.choices[0].message.content.strip()
    except Exception as e:
        logger.error(f"Error getting LLM summary: {e}")
        return f"The agent used tool '{tool_name}' and got a result."

async def _prepare_messages_for_llm(full_history: List[Any]) -> List[Any]:
    last_tool_result_idx = -1
    for i in range(len(full_history) - 1, -1, -1):
        msg = full_history[i]
        role = msg.get('role') if isinstance(msg, dict) else getattr(msg, 'role', None)
        if role == 'tool':
            last_tool_result_idx = i
            break
    if last_tool_result_idx == -1:
        return full_history

    user_prompt = ""
    for msg in full_history:
        if (isinstance(msg, dict) and msg.get('role') == 'user') or \
           (not isinstance(msg, dict) and getattr(msg, 'role', None) == 'user'):
            user_prompt = msg.get('content') if isinstance(msg, dict) else getattr(msg, 'content', '')
            break

    turn_summaries = []
    i = 0
    while i < last_tool_result_idx:
        msg = full_history[i]
        role = msg.get('role') if isinstance(msg, dict) else getattr(msg, 'role', None)
        if role != 'assistant':
            i += 1
            continue

        tool_calls = msg.get('tool_calls') if isinstance(msg, dict) else getattr(msg, 'tool_calls', None)
        if not tool_calls:
            i += 1
            continue
            
        context_for_llm = user_prompt + "\n" + "\n".join(turn_summaries)
        num_tools_in_call = len(tool_calls)
        tool_results_msgs = full_history[i + 1 : i + 1 + num_tools_in_call]
        
        for tool_call, tool_result_msg in zip(tool_calls, tool_results_msgs):
            res_role = tool_result_msg.get('role') if isinstance(tool_result_msg, dict) else getattr(tool_result_msg, 'role', None)
            if res_role != 'tool':
                break
            
            tool_name = getattr(tool_call.function, 'name', 'unknown_tool')
            tool_result_content = tool_result_msg.get('content', '')
            turn_summary = await _get_llm_summary_of_turn(context_for_llm, tool_name, tool_result_content)
            turn_summaries.append(turn_summary)

        i += 1 + num_tools_in_call
    
    final_messages = []
    for msg in full_history:
        role = msg.get('role') if isinstance(msg, dict) else getattr(msg, 'role', None)
        if role in ['system', 'user']:
            final_messages.append(msg)
        if role == 'user':
            break

    if turn_summaries:
        summary_context_message = {
            "role": "system",
            "content": "[Summary of Previous Actions]:\n" + "\n".join(f"- {s}" for s in turn_summaries)
        }
        final_messages.append(summary_context_message)
        
    final_messages.extend(full_history[last_tool_result_idx - len(tool_calls):])
    return final_messages



def fix_openapi_schema_recursive(schema_node: Any, path: str = "") -> Any:
    if not isinstance(schema_node, dict):
        return schema_node
    if schema_node.get("type") == "array" and (schema_node.get("items") is None or (isinstance(schema_node.get("items"), dict) and not schema_node.get("items"))):
        logger.warning(f"Schema Fix: Adding default 'items': {{'type': 'string'}} for array at path '{path}'.")
        schema_node["items"] = {"type": "string"}
    for key, value in schema_node.items():
        new_path = f"{path}.{key}" if path else key
        if isinstance(value, dict):
            fix_openapi_schema_recursive(value, new_path)
        elif isinstance(value, list):
            for i, item_in_list in enumerate(value):
                if isinstance(item_in_list, dict):
                    fix_openapi_schema_recursive(item_in_list, f"{new_path}[{i}]")
    return schema_node

async def get_tools_for_mcp(session: ClientSession, category_prefix: str):
    logger.info(f"Fetching tools from MCP '{category_prefix}'...")
    try:
        resp = await session.list_tools()
        logger.info(f"Found {len(resp.tools)} tools for '{category_prefix}'.")
        oa_tools = []
        for t in resp.tools:
            if hasattr(t, 'name') and hasattr(t, 'inputSchema'):
                prefixed_tool_name = f"{category_prefix}_{t.name}"
                parameters_schema = fix_openapi_schema_recursive(copy.deepcopy(t.inputSchema or {"type": "object", "properties": {}}), prefixed_tool_name)
                if "type" not in parameters_schema: parameters_schema["type"] = "object"
                if parameters_schema["type"] == "object" and "properties" not in parameters_schema: parameters_schema["properties"] = {}
                
                oa_tools.append({
                    "type": "function", "function": {
                        "name": prefixed_tool_name,
                        "description": f"({category_prefix}) {t.description or 'Tool from '+category_prefix}",
                        "parameters": parameters_schema,
                    },
                })
            else:
                logger.warning(f"Skipping tool from '{category_prefix}' due to missing attributes: {t}")
        return oa_tools
    except Exception as e:
        logger.error(f"Failed to get tools from MCP '{category_prefix}': {e}", exc_info=True)
        return []

async def execute_tool_call(mcp_sessions: dict[str, ClientSession], tool_call, test_context: Optional[Dict] = None, websocket: Optional[WebSocket] = None, run_id: Optional[str] = None):
    prefixed_tool_name = tool_call.function.name
    tool_call_id = tool_call.id
    logger.info(f"Attempting Tool Call: {prefixed_tool_name}")
    tool_output = "[TOOL ERROR] Unknown execution issue"
    try:
        category, original_tool_name = prefixed_tool_name.split("_", 1)
    except ValueError:
        tool_output = f"[TOOL ERROR] Invalid tool name format. Expected 'category_toolname', got '{prefixed_tool_name}'."
        if debug_logger.hasHandlers():
            debug_logger.info("\n" + "="*20 + " TOOL OUTPUT " + "="*20)
            debug_logger.info(tool_output)
        return {"role": "tool", "tool_call_id": tool_call_id, "content": tool_output}

    if category not in mcp_sessions:
        tool_output = f"[TOOL ERROR] MCP category '{category}' is not available or not connected."
        if debug_logger.hasHandlers():
            debug_logger.info("\n" + "="*20 + " TOOL OUTPUT " + "="*20)
            debug_logger.info(tool_output)
        return {"role": "tool", "tool_call_id": tool_call_id, "content": tool_output}
    
    session = mcp_sessions[category]
    tool_args = {}
    try:
        tool_args = json.loads(tool_call.function.arguments or "{}")
        
        if debug_logger.hasHandlers():
            debug_logger.info("\n" + "="*20 + " TOOL CALL " + "="*20)
            debug_logger.info(f"Tool: {prefixed_tool_name}")
            debug_logger.info(f"Arguments:\n{json.dumps(tool_args, indent=2)}")

        if category == "browser" and prefixed_tool_name in ACTIONS_REQUIRING_ELEMENT_INTERACTION and prefixed_tool_name != "browser_hover" and tool_args.get("mcp_id"):
            try:
                await session.call_tool("hover", {"mcp_id": tool_args["mcp_id"]})
            except Exception as hover_exc:
                logger.warning(f"Pre-hover for {tool_args['mcp_id']} failed: {hover_exc}. Proceeding.")
        
        tool_exec_result = await session.call_tool(original_tool_name, tool_args)

        if (original_tool_name == 'browser_screenshot' or original_tool_name == 'browser_take_screenshot') and run_id and websocket:
            logging.info("Handling browser_screenshot result for web UI...")
            try:
                content_list = getattr(tool_exec_result, 'content', None)
                screenshot_data_base64 = None
                if isinstance(content_list, list) :
                    for item in content_list:
                        if hasattr(item, 'type') and item.type == 'image' and hasattr(item, 'data'):
                            screenshot_data_base64 = item.data
                            break 
                    
                if screenshot_data_base64:
                    image_bytes = base64.b64decode(screenshot_data_base64)
                    screenshot_dir = os.path.join('results', run_id, 'screenshots')
                    os.makedirs(screenshot_dir, exist_ok=True)
                    
                    file_extension = 'png' # default
                    if hasattr(content_list[0], 'mimeType'):
                        mime_type = content_list[0].mimeType
                        ext = mime_type.split('/')[-1]
                        if ext: file_extension = 'jpg' if ext == 'jpeg' else ext

                    s_no = test_context.get("s_no", "NA")
                    step_index = test_context.get("step_index", "NA")
                    step_part = f"step{step_index}" if isinstance(step_index, int) else str(step_index)
                    base_filename = f"Screenshot-testcase{s_no}-{step_part}"
                    filename = os.path.join(screenshot_dir, f"{base_filename}.{file_extension}")

                    with open(filename, 'wb') as f: f.write(image_bytes)
                    
                    web_path = f"/results/{run_id}/screenshots/{base_filename}.{file_extension}"
                    await websocket.send_json({
                        "type": "screenshot",
                        "path": web_path,
                        "s_no": test_context.get("s_no", "N/A"),
                        "step": test_context.get("step_index", "N/A")
                    })
                    tool_output_str = json.dumps({"status": "success", "message": f"Screenshot sent to client and saved to {filename}"})
                else:
                    tool_output_str = json.dumps({"status": "warning", "message": "Screenshot executed, but no image data found."})
            except Exception as screenshot_err:
                tool_output_str = json.dumps({"status": "error", "message": f"Failed to process screenshot: {screenshot_err}"})
        else:
            tool_output_obj = tool_exec_result.model_dump()
            tool_output_str = json.dumps(tool_output_obj, ensure_ascii=False, default=str)
        
        tool_output = tool_output_str

    except Exception as tool_exc:
        error_message = getattr(tool_exc, 'message', str(tool_exc))
        if hasattr(tool_exc, 'details'): error_message += f" Details: {tool_exc.details}"
        logger.error(f"Error executing tool '{prefixed_tool_name}': {error_message}", exc_info=True)
        tool_output = f"[TOOL ERROR] Execution failed for {prefixed_tool_name}: {error_message}"
    
    if debug_logger.hasHandlers():
        debug_logger.info("\n" + "="*20 + " TOOL OUTPUT " + "="*20)
        debug_logger.info(tool_output)

    return {"role": "tool", "tool_call_id": tool_call_id, "content": tool_output}

async def process_user_request(user_input: str, mcp_sessions: dict[str, ClientSession], all_oa_tools: list, test_context: Optional[Dict] = None, websocket: Optional[WebSocket] = None, run_id: Optional[str] = None) -> str:
    logger.info(f"Processing user request: '{user_input}'")
    messages: list[Any] = [{"role": "system", "content": SYSTEM_PROMPT}, {"role": "user", "content": user_input}]
    turn_count = 0
    while turn_count < MAX_TURNS:
        turn_count += 1
        logger.info(f"--- Turn {turn_count}/{MAX_TURNS} ---")
        
        messages_for_api_call = copy.deepcopy(messages)

        tool_id_to_name_map = {}
        for msg in messages_for_api_call:
            if msg.get("role") == "assistant" and msg.get("tool_calls"):
                for tc in msg["tool_calls"]:
                    if tc.get("id") and tc.get("function", {}).get("name"):
                        tool_id_to_name_map[tc["id"]] = tc["function"]["name"]

        last_non_tool_msg_index = -1
        for i in range(len(messages_for_api_call) - 1, -1, -1):
            if messages_for_api_call[i].get("role") != "tool":
                last_non_tool_msg_index = i
                break

        if last_non_tool_msg_index != -1:
            for i in range(last_non_tool_msg_index + 1):
                msg = messages_for_api_call[i]
                if msg.get("role") == "tool":
                    tool_call_id = msg.get("tool_call_id")
                    tool_name = tool_id_to_name_map.get(tool_call_id)
                    
                    if tool_name and tool_name.startswith("browser_"):
                        content = msg.get("content")
                        if isinstance(content, str) and len(content) > OLD_BROWSER_SNAPSHOT_TRUNCATION_LENGTH:
                            msg["content"] = content[:OLD_BROWSER_SNAPSHOT_TRUNCATION_LENGTH] + f"\n[... content truncated ...]"


        if debug_logger.hasHandlers():
            debug_logger.info("\n" + "#"*20 + " LLM INPUT " + "#"*20)
            debug_logger.info(f"--- Turn {turn_count}/{MAX_TURNS} ---")
            try:
                debug_logger.info(json.dumps(messages_for_api_call, indent=2))
            except TypeError: 
                debug_logger.info(str(messages_for_api_call))
                
        try:
            response_message = None
            client = get_openai_client()
            response = await client.chat.completions.create(model=OPENAI_MODEL, messages=messages_for_api_call, tools=all_oa_tools, tool_choice="auto")
            response_message = response.choices[0].message
            
            response_message_dict = response_message.model_dump(exclude_none=True)
            if debug_logger.hasHandlers():
                debug_logger.info("\n" + "#"*20 + " LLM OUTPUT " + "#"*20)
                debug_logger.info(json.dumps(response_message_dict, indent=2))
            messages.append(response_message_dict)

            if response_message and response_message.tool_calls:
                tool_results = []
                for tool_call in response_message.tool_calls:
                    result = await execute_tool_call(mcp_sessions, tool_call, test_context=test_context, websocket=websocket, run_id=run_id)
                    tool_results.append(result)
                messages.extend(tool_results)
                continue
            else:
                final_content = response_message.content if response_message else "[No response message]"
                if run_logger.hasHandlers(): run_logger.info(f"Assistant: {final_content.strip()}")
                if debug_logger.hasHandlers():
                    debug_logger.info("\n" + "*"*20 + " FINAL AGENT MESSAGE " + "*"*20)
                    debug_logger.info(final_content.strip())
                return final_content
        except Exception as e:
            logger.error(f"Error in processing loop turn {turn_count}: {e}", exc_info=True)
            return f"[ERROR] An unexpected error occurred: {e}"

    logger.warning(f"Reached max turns ({MAX_TURNS}).")
    return "[MAX_TURNS_REACHED_NO_FINAL_MESSAGE]"

def apply_failure_styling(output_xlsx_path: str, status_column_name: str = "Status"):
    try:
        workbook = load_workbook(output_xlsx_path)
        sheet = workbook.active
        status_col_idx = None
        for col_idx, cell in enumerate(sheet[1], 1):
            if cell.value == status_column_name:
                status_col_idx = col_idx
                break
        if status_col_idx is None: return

        red_font = Font(color="FF0000")
        for row_idx in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row_idx, column=status_col_idx)
            if cell.value and isinstance(cell.value, str) and cell.value.lower().startswith("failed"):
                cell.font = red_font
        workbook.save(output_xlsx_path)
    except Exception as e:
        logger.error(f"Error applying styling to '{output_xlsx_path}': {e}", exc_info=True)

async def _execute_one_test_case(
    test_case_row: pd.Series, test_case_index: int, df: pd.DataFrame,
    mcp_sessions: Dict[str, ClientSession], all_oa_tools: List[Dict],
    websocket: WebSocket, run_id: str
) -> bool:
    """Executes a single test case and returns True for success, False for failure."""
    await initialize_mcp_manager()
    s_no = test_case_row.get("S.No", f"Row {test_case_index+2}")
    heading = str(test_case_row.get("Test case heading", "N/A"))
    
    # Log test case start to the shared debug log
    if debug_logger.hasHandlers():
        debug_logger.info("\n" + "="*40 + f" START TEST CASE S.No: {s_no} " + "="*40)
        debug_logger.info(f"Heading: {heading}")
        debug_logger.info(f"Timestamp: {datetime.datetime.now().isoformat()}")
        debug_logger.info("="*105)

    steps_str = str(test_case_row.get("Steps", ""))
    verification_str = str(test_case_row.get("Verification", ""))

    start_message = f"▶️ Processing Test Case S.No: {s_no} - Heading: {heading}"
    await websocket.send_json({"type": "log", "level": "info", "message": start_message, "s_no": s_no, "is_header": True})
    run_logger.info("=" * 80 + f"\n{start_message}")

    overall_success = True
    failure_details = []

    if steps_str.strip():
        step_commands = [s.strip() for s in steps_str.splitlines() if s.strip()]
        for i, step_cmd in enumerate(step_commands):
            step_message = f"  Executing Step {i+1}/{len(step_commands)}: {step_cmd[:100]}{'...' if len(step_cmd)>100 else ''}"
            await websocket.send_json({"type": "log", "level": "info", "message": step_message, "s_no": s_no})
            run_logger.info(f"--- Step {i+1}/{len(step_commands)} ---\nInstruction: {step_cmd}")
            
            if debug_logger.hasHandlers():
                debug_logger.info("\n" + "*"*20 + " USER INPUT (STEP) " + "*"*20)
                debug_logger.info(f"Step {i+1}: {step_cmd}")

            test_context = {"s_no": s_no, "step_index": i + 1}
            step_response = await process_user_request(step_cmd, mcp_sessions, all_oa_tools, test_context=test_context, websocket=websocket, run_id=run_id)

            if any(err in step_response for err in ["[TOOL ERROR]", "[ERROR]", "[MAX_TURNS_REACHED"]):
                overall_success = False
                failure_summary = step_response[:250] + "..." if len(step_response) > 250 else step_response
                failure_details.append(f"Step {i+1} failed: {failure_summary}")

    if verification_str.strip():
        verif_message = f"  Performing Verification: {verification_str[:100]}{'...' if len(verification_str)>100 else ''}"
        await websocket.send_json({"type": "log", "level": "info", "message": verif_message, "s_no": s_no})
        run_logger.info(f"--- Verification ---\nInstruction: {verification_str}")
        
        verification_prompt = (f"{verification_str.strip()}\n\nBased on your actions, conclude with 'VERIFICATION_SUCCESSFUL.' or 'VERIFICATION_FAILED: [reason].'")
        
        if debug_logger.hasHandlers():
            debug_logger.info("\n" + "*"*20 + " USER INPUT (VERIFICATION) " + "*"*20)
            debug_logger.info(f"Verification: {verification_str.strip()}")
        
        verif_context = {"s_no": s_no, "step_index": "verification"}
        verif_output = await process_user_request(verification_prompt, mcp_sessions, all_oa_tools, test_context=verif_context, websocket=websocket, run_id=run_id)

        if "VERIFICATION_SUCCESSFUL." not in verif_output:
            overall_success = False
            reason = verif_output.split("VERIFICATION_FAILED:", 1)[-1].strip() if "VERIFICATION_FAILED:" in verif_output else verif_output
            failure_details.append(f"Verification failed: {reason}")
    
    final_status = "Success"
    if not overall_success:
        final_status = f"Failed: {'; '.join(failure_details)}"
    
    if debug_logger.hasHandlers():
        debug_logger.info("\n" + "="*40 + f" END TEST CASE S.No: {s_no} " + "="*40)
        debug_logger.info(f"Final Status: {final_status}")
        debug_logger.info("="*105 + "\n")

    df.loc[test_case_index, 'Status'] = final_status
    result_message = f"--- Test Case S.No: {s_no} Result: {final_status.split(':',1)[0]} ---"
    await websocket.send_json({"type": "log", "level": "result", "message": result_message, "s_no": s_no, "status": final_status})
    run_logger.info(f"--- Test Case S.No: {s_no} Result ---\nFinal Status: {final_status}")
    
    return overall_success

async def run_test_suite_from_web(xlsx_file_path: str, websocket: WebSocket, run_id: str):
    global LLM_PROVIDER, SCREENSHOT_DIR
    
    debug_handler = None
    try:
        # --- Setup Debug Logger for the entire run based on Excel filename ---
        os.makedirs('debuglogs', exist_ok=True)
        input_basename = os.path.basename(xlsx_file_path)
        base_name, _ = os.path.splitext(input_basename)
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        debug_log_filename = f"{base_name}_{timestamp}_debug.log"
        debug_log_path = os.path.join('debuglogs', debug_log_filename)
        
        debug_handler = logging.FileHandler(debug_log_path, mode='w', encoding='utf-8')
        debug_handler.setFormatter(logging.Formatter('%(message)s'))
        if debug_logger.hasHandlers(): debug_logger.handlers.clear()
        debug_logger.addHandler(debug_handler)
        await websocket.send_json({"type": "log", "level": "system", "message": f"📝 Detailed debug log for this run: '{debug_log_filename}'"})
        # --- End Setup ---

        load_dotenv()
        LLM_PROVIDER = os.getenv("LLM_PROVIDER", "openai").lower()
        if LLM_PROVIDER == 'google':
            genai.configure(api_key=os.getenv("GOOGLE_API_KEY"))
        else:
            openai.api_key = os.getenv("OPENAI_API_KEY")
            openai.base_url = os.getenv("BASE_URL")
        
        results_dir = os.path.join('results', run_id)
        os.makedirs(results_dir, exist_ok=True)
        SCREENSHOT_DIR = os.path.join(results_dir, 'screenshots')
        os.makedirs(SCREENSHOT_DIR, exist_ok=True)
        
        log_file_path = os.path.join(results_dir, "run_log.txt")
        if run_logger.hasHandlers(): run_logger.handlers.clear()
        file_handler = logging.FileHandler(log_file_path, mode='w', encoding='utf-8')
        file_handler.setFormatter(logging.Formatter('%(asctime)s - %(message)s'))
        run_logger.addHandler(file_handler)
        
        await websocket.send_json({"type": "log", "level": "system", "message": f"📂 Results will be saved on server in: {os.path.abspath(results_dir)}"})
        
        mcp_sessions: dict[str, ClientSession] = {}
        all_oa_tools: list = []
        async with AsyncExitStack() as stack:
            for category, url in MCP_ENDPOINTS.items():
                try:
                    sse_mgr = sse_client(url)
                    rx, tx = await stack.enter_async_context(sse_mgr)
                    session_mgr = ClientSession(rx, tx)
                    sess = await stack.enter_async_context(session_mgr)
                    await sess.initialize()
                    mcp_sessions[category] = sess
                    oa_tools_for_mcp = await get_tools_for_mcp(sess, category)
                    all_oa_tools.extend(oa_tools_for_mcp)
                except Exception as e:
                    logger.error(f"Failed to connect to MCP '{category}': {e}")
            
            if not mcp_sessions:
                await websocket.send_json({"type": "log", "level": "error", "message": "❌ Error: No MCP instances available. Check mcp-proxy setups."})
                return

            df = pd.read_excel(xlsx_file_path)
            if 'Status' not in df.columns: df['Status'] = pd.Series(dtype='str')

            mandatory_column_name = None
            possible_names = ['Mandatory', 'Execution Type', 'Type', 'Pass Status']
            for name in possible_names:
                if name in df.columns:
                    mandatory_column_name = name
                    log_msg = f"ℹ️ Using column '{mandatory_column_name}' to check for mandatory test cases (e.g., value 'M')."
                    await websocket.send_json({"type": "log", "level": "system", "message": log_msg})
                    run_logger.info(log_msg)
                    break

            for index, row_data in df.iterrows():
                test_passed = await _execute_one_test_case(row_data, index, df, mcp_sessions, all_oa_tools, websocket, run_id)

                if not test_passed:
                    is_mandatory = False
                    if mandatory_column_name:
                        mandatory_flag = str(row_data.get(mandatory_column_name, "")).strip().upper()
                        if mandatory_flag in ['M', 'Y', 'YES', 'MANDATORY']:
                            is_mandatory = True
                    
                    if is_mandatory:
                        s_no = row_data.get("S.No", f"Row {index+2}")
                        stop_message = f"🛑 Halting test suite: Mandatory test case (S.No: {s_no}) failed."
                        await websocket.send_json({"type": "log", "level": "error", "message": stop_message, "s_no": s_no})
                        run_logger.error(stop_message)
                        break

        input_basename = os.path.basename(xlsx_file_path)
        base, ext = os.path.splitext(input_basename)
        output_filename = f"{base}_results{ext}"
        final_output_path = os.path.join(results_dir, output_filename)
        
        df.to_excel(final_output_path, index=False, engine='openpyxl')
        apply_failure_styling(final_output_path)

        web_result_path = f"/results/{run_id}/{output_filename}"
        await websocket.send_json({"type": "result_file", "path": web_result_path, "filename": output_filename})
        run_logger.info(f"--- Test Run Complete. Results saved to {final_output_path} ---")

    except Exception as e:
        error_msg = f"A critical error occurred: {e}"
        logger.error(error_msg, exc_info=True)
        if run_logger.hasHandlers(): run_logger.error(error_msg + f"\n{traceback.format_exc()}")
        if debug_logger.hasHandlers(): debug_logger.error("CRITICAL ERROR\n" + traceback.format_exc())
        await websocket.send_json({"type": "log", "level": "error", "message": f"❌ CRITICAL ERROR: {error_msg}."})
    finally:
        # --- Teardown Debug Logger ---
        if debug_handler:
            debug_handler.close()
            debug_logger.removeHandler(debug_handler)
        # --- End Teardown ---